"""
Système de Génération de Rapports pour Robot de Trading Algorithmique IA
========================================================================

Ce module génère des rapports complets et professionnels sur tous les aspects
du système de trading : performance, risques, exécution, système, et compliance.
Support pour multiple formats et distribution automatique.

Fonctionnalités:
- Rapports de performance détaillés (P&L, métriques, attribution)
- Analyse des trades et stratégies
- Rapports de risque et exposition
- Monitoring système et infrastructure
- Export multi-format (PDF, HTML, Excel, JSON)
- Dashboards interactifs temps réel
- Distribution automatique (email, S3, webhook)
- Templates personnalisables
- Rapports de compliance réglementaire

Auteur: Robot Trading IA System
Version: 1.0.0
Date: 2025
"""

import asyncio
import json
import os
import sys
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Union
import io
import base64
from dataclasses import dataclass, field
from enum import Enum
import hashlib
import tempfile
import shutil

# Data analysis
import pandas as pd
import numpy as np
from scipy import stats

# Visualization
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# Document generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, Image, PageBreak, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.barcharts import VerticalBarChart

# Excel generation
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# Email
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Cloud storage
import boto3
from azure.storage.blob import BlobServiceClient

# Web frameworks for dashboards
from flask import Flask, render_template_string
import dash
from dash import dcc, html, dash_table
from dash.dependencies import Input, Output
import dash_bootstrap_components as dbc

# Logging and monitoring
from rich.console import Console
from rich.table import Table as RichTable
from rich.panel import Panel
from rich.progress import track

# Local imports
sys.path.append(str(Path(__file__).parent.parent))
from utils.logger import get_structured_logger
from utils.metrics import MetricsCollector
from config import get_config

console = Console()
logger = get_structured_logger(__name__)


class ReportType(Enum):
    """Types de rapports disponibles"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ANNUAL = "annual"
    REAL_TIME = "real_time"
    ON_DEMAND = "on_demand"
    REGULATORY = "regulatory"
    RISK = "risk"
    PERFORMANCE = "performance"
    SYSTEM = "system"
    COMPLIANCE = "compliance"


class ReportFormat(Enum):
    """Formats de sortie des rapports"""
    PDF = "pdf"
    HTML = "html"
    EXCEL = "xlsx"
    CSV = "csv"
    JSON = "json"
    MARKDOWN = "md"
    DASHBOARD = "dashboard"


class DeliveryMethod(Enum):
    """Méthodes de livraison des rapports"""
    EMAIL = "email"
    S3 = "s3"
    AZURE_BLOB = "azure_blob"
    WEBHOOK = "webhook"
    FTP = "ftp"
    LOCAL = "local"
    API = "api"


@dataclass
class ReportConfig:
    """Configuration d'un rapport"""
    name: str
    type: ReportType
    format: ReportFormat
    schedule: Optional[str] = None  # Cron expression
    recipients: List[str] = field(default_factory=list)
    delivery_methods: List[DeliveryMethod] = field(default_factory=list)
    filters: Dict[str, Any] = field(default_factory=dict)
    template: Optional[str] = None
    include_sections: List[str] = field(default_factory=list)
    custom_params: Dict[str, Any] = field(default_factory=dict)


@dataclass
class PerformanceMetrics:
    """Métriques de performance pour les rapports"""
    total_pnl: Decimal
    realized_pnl: Decimal
    unrealized_pnl: Decimal
    return_pct: float
    sharpe_ratio: float
    sortino_ratio: float
    calmar_ratio: float
    max_drawdown: float
    max_drawdown_duration: int
    win_rate: float
    profit_factor: float
    expectancy: float
    trades_count: int
    winning_trades: int
    losing_trades: int
    avg_win: Decimal
    avg_loss: Decimal
    largest_win: Decimal
    largest_loss: Decimal
    consecutive_wins: int
    consecutive_losses: int
    recovery_factor: float
    payoff_ratio: float
    volatility: float
    beta: float
    alpha: float
    correlation_benchmark: float


class ReportGenerator:
    """Générateur principal de rapports"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        self.config = config or get_config()
        self.reports_dir = Path(self.config.get('reports_dir', './reports'))
        self.templates_dir = Path(self.config.get('templates_dir', './templates'))
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        
        # Cache pour les données
        self._data_cache = {}
        self._cache_ttl = 300  # 5 minutes
        
        # Styles pour les rapports
        self._init_styles()
        
        # Métriques
        self.metrics_collector = MetricsCollector()
        
        logger.info("Report generator initialized", 
                   reports_dir=str(self.reports_dir))
    
    def _init_styles(self):
        """Initialise les styles pour les rapports"""
        # Matplotlib style
        plt.style.use('seaborn-v0_8-darkgrid')
        sns.set_palette("husl")
        
        # Couleurs corporate
        self.colors = {
            'primary': '#1f77b4',
            'secondary': '#ff7f0e',
            'success': '#2ca02c',
            'danger': '#d62728',
            'warning': '#ff9800',
            'info': '#17a2b8',
            'light': '#f8f9fa',
            'dark': '#343a40'
        }
        
        # Styles PDF
        self.pdf_styles = getSampleStyleSheet()
        self.pdf_styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.pdf_styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(self.colors['primary']),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        self.pdf_styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=self.pdf_styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor(self.colors['dark']),
            spaceAfter=12
        ))
    
    async def generate_report(self, config: ReportConfig) -> Path:
        """
        Génère un rapport selon la configuration
        
        Args:
            config: Configuration du rapport
            
        Returns:
            Path du rapport généré
        """
        logger.info(f"Generating report: {config.name}", 
                   type=config.type.value,
                   format=config.format.value)
        
        try:
            # Collecter les données
            data = await self._collect_report_data(config)
            
            # Générer selon le format
            if config.format == ReportFormat.PDF:
                report_path = await self._generate_pdf_report(config, data)
            elif config.format == ReportFormat.HTML:
                report_path = await self._generate_html_report(config, data)
            elif config.format == ReportFormat.EXCEL:
                report_path = await self._generate_excel_report(config, data)
            elif config.format == ReportFormat.JSON:
                report_path = await self._generate_json_report(config, data)
            elif config.format == ReportFormat.DASHBOARD:
                report_path = await self._generate_dashboard(config, data)
            else:
                raise ValueError(f"Unsupported format: {config.format}")
            
            # Livrer le rapport
            await self._deliver_report(config, report_path)
            
            logger.info(f"Report generated successfully: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"Report generation failed: {str(e)}", exc_info=True)
            raise
    
    async def _collect_report_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Collecte les données pour le rapport"""
        cache_key = self._get_cache_key(config)
        
        # Vérifier le cache
        if cache_key in self._data_cache:
            cached_data, timestamp = self._data_cache[cache_key]
            if datetime.now().timestamp() - timestamp < self._cache_ttl:
                logger.debug("Using cached data for report")
                return cached_data
        
        logger.info("Collecting data for report")
        
        # Déterminer la période
        end_date = datetime.now()
        if config.type == ReportType.DAILY:
            start_date = end_date - timedelta(days=1)
        elif config.type == ReportType.WEEKLY:
            start_date = end_date - timedelta(weeks=1)
        elif config.type == ReportType.MONTHLY:
            start_date = end_date - timedelta(days=30)
        elif config.type == ReportType.QUARTERLY:
            start_date = end_date - timedelta(days=90)
        elif config.type == ReportType.ANNUAL:
            start_date = end_date - timedelta(days=365)
        else:
            start_date = config.filters.get('start_date', end_date - timedelta(days=30))
        
        # Collecter les données selon les sections
        data = {
            'metadata': {
                'report_name': config.name,
                'report_type': config.type.value,
                'generated_at': datetime.now().isoformat(),
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                }
            }
        }
        
        # Performance metrics
        if 'performance' in config.include_sections or not config.include_sections:
            data['performance'] = await self._collect_performance_data(start_date, end_date)
        
        # Trading activity
        if 'trading' in config.include_sections or not config.include_sections:
            data['trading'] = await self._collect_trading_data(start_date, end_date)
        
        # Risk metrics
        if 'risk' in config.include_sections or not config.include_sections:
            data['risk'] = await self._collect_risk_data(start_date, end_date)
        
        # System metrics
        if 'system' in config.include_sections or not config.include_sections:
            data['system'] = await self._collect_system_data(start_date, end_date)
        
        # Market data
        if 'market' in config.include_sections:
            data['market'] = await self._collect_market_data(start_date, end_date)
        
        # Cache les données
        self._data_cache[cache_key] = (data, datetime.now().timestamp())
        
        return data
    
    async def _collect_performance_data(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Collecte les données de performance"""
        # Simuler la collecte de données pour la démo
        # En production, interroger la base de données
        
        # Générer des données de P&L quotidien
        days = (end_date - start_date).days
        dates = pd.date_range(start=start_date, end=end_date, freq='D')
        
        # Simuler les données
        np.random.seed(42)
        daily_returns = np.random.normal(0.001, 0.02, days)
        cumulative_returns = (1 + daily_returns).cumprod() - 1
        
        # Calculer les métriques
        total_return = cumulative_returns[-1]
        volatility = daily_returns.std() * np.sqrt(252)
        sharpe_ratio = (daily_returns.mean() * 252) / volatility if volatility > 0 else 0
        
        # Drawdown
        running_max = np.maximum.accumulate(1 + cumulative_returns)
        drawdown = (1 + cumulative_returns) / running_max - 1
        max_drawdown = drawdown.min()
        
        # Trades simulés
        n_trades = np.random.randint(50, 200)
        win_rate = np.random.uniform(0.45, 0.65)
        
        return {
            'summary': {
                'total_return': float(total_return),
                'total_pnl': float(total_return * 100000),  # Sur 100k de capital
                'sharpe_ratio': float(sharpe_ratio),
                'volatility': float(volatility),
                'max_drawdown': float(max_drawdown),
                'win_rate': float(win_rate),
                'total_trades': n_trades
            },
            'daily_pnl': {
                'dates': dates.tolist(),
                'pnl': (daily_returns * 100000).tolist(),
                'cumulative_pnl': (cumulative_returns * 100000).tolist()
            },
            'metrics': self._calculate_performance_metrics(daily_returns, n_trades, win_rate)
        }
    
    def _calculate_performance_metrics(self, returns: np.ndarray, n_trades: int, win_rate: float) -> PerformanceMetrics:
        """Calcule les métriques de performance détaillées"""
        # Simuler les métriques pour la démo
        winning_trades = int(n_trades * win_rate)
        losing_trades = n_trades - winning_trades
        
        avg_win = abs(np.random.normal(500, 200))
        avg_loss = abs(np.random.normal(300, 150))
        
        return PerformanceMetrics(
            total_pnl=Decimal(str(returns.sum() * 100000)),
            realized_pnl=Decimal(str(returns.sum() * 100000 * 0.8)),
            unrealized_pnl=Decimal(str(returns.sum() * 100000 * 0.2)),
            return_pct=float(returns.sum()),
            sharpe_ratio=float((returns.mean() * 252) / (returns.std() * np.sqrt(252))),
            sortino_ratio=float((returns.mean() * 252) / (returns[returns < 0].std() * np.sqrt(252))),
            calmar_ratio=float(returns.mean() * 252 / abs(returns.min())),
            max_drawdown=float(self._calculate_max_drawdown(returns)),
            max_drawdown_duration=30,
            win_rate=win_rate,
            profit_factor=float((winning_trades * avg_win) / (losing_trades * avg_loss)) if losing_trades > 0 else 0,
            expectancy=float((win_rate * avg_win) - ((1 - win_rate) * avg_loss)),
            trades_count=n_trades,
            winning_trades=winning_trades,
            losing_trades=losing_trades,
            avg_win=Decimal(str(avg_win)),
            avg_loss=Decimal(str(avg_loss)),
            largest_win=Decimal(str(avg_win * 3)),
            largest_loss=Decimal(str(avg_loss * 2.5)),
            consecutive_wins=5,
            consecutive_losses=3,
            recovery_factor=2.5,
            payoff_ratio=float(avg_win / avg_loss) if avg_loss > 0 else 0,
            volatility=float(returns.std() * np.sqrt(252)),
            beta=0.8,
            alpha=0.02,
            correlation_benchmark=0.65
        )
    
    def _calculate_max_drawdown(self, returns: np.ndarray) -> float:
        """Calcule le drawdown maximum"""
        cumulative = (1 + returns).cumprod()
        running_max = np.maximum.accumulate(cumulative)
        drawdown = (cumulative - running_max) / running_max
        return float(drawdown.min())
    
    async def _collect_trading_data(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Collecte les données de trading"""
        # Simuler pour la démo
        strategies = ['Statistical Arbitrage', 'Market Making', 'Scalping']
        symbols = ['BTC-USDT', 'ETH-USDT', 'BNB-USDT']
        
        trades = []
        for i in range(100):
            trade_date = start_date + timedelta(
                seconds=np.random.randint(0, int((end_date - start_date).total_seconds()))
            )
            
            trades.append({
                'id': f'TRADE-{i+1:04d}',
                'timestamp': trade_date.isoformat(),
                'symbol': np.random.choice(symbols),
                'strategy': np.random.choice(strategies),
                'side': np.random.choice(['BUY', 'SELL']),
                'quantity': float(np.random.uniform(0.01, 1.0)),
                'price': float(np.random.uniform(30000, 60000)),
                'pnl': float(np.random.normal(0, 500)),
                'commission': float(np.random.uniform(1, 10))
            })
        
        # Analyser par stratégie
        strategy_performance = {}
        for strategy in strategies:
            strategy_trades = [t for t in trades if t['strategy'] == strategy]
            total_pnl = sum(t['pnl'] for t in strategy_trades)
            
            strategy_performance[strategy] = {
                'trades_count': len(strategy_trades),
                'total_pnl': total_pnl,
                'avg_pnl': total_pnl / len(strategy_trades) if strategy_trades else 0,
                'win_rate': len([t for t in strategy_trades if t['pnl'] > 0]) / len(strategy_trades) if strategy_trades else 0
            }
        
        return {
            'summary': {
                'total_trades': len(trades),
                'total_volume': sum(t['quantity'] * t['price'] for t in trades),
                'total_commission': sum(t['commission'] for t in trades),
                'avg_trade_size': np.mean([t['quantity'] * t['price'] for t in trades])
            },
            'trades': trades[:20],  # Derniers 20 trades
            'strategy_performance': strategy_performance,
            'symbol_distribution': {
                symbol: len([t for t in trades if t['symbol'] == symbol])
                for symbol in symbols
            }
        }
    
    async def _collect_risk_data(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Collecte les données de risque"""
        # Simuler pour la démo
        return {
            'var': {
                '95': -5000,  # VaR 95%
                '99': -8000   # VaR 99%
            },
            'exposure': {
                'BTC-USDT': 0.35,
                'ETH-USDT': 0.30,
                'BNB-USDT': 0.20,
                'CASH': 0.15
            },
            'correlations': {
                'BTC-ETH': 0.85,
                'BTC-BNB': 0.75,
                'ETH-BNB': 0.80
            },
            'leverage': {
                'current': 1.5,
                'average': 1.8,
                'maximum': 2.5
            },
            'stress_tests': {
                'market_crash_20pct': -15000,
                'flash_crash': -8000,
                'black_swan': -25000
            }
        }
    
    async def _collect_system_data(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Collecte les données système"""
        # Simuler pour la démo
        return {
            'uptime': {
                'percentage': 99.95,
                'total_hours': 720,
                'downtime_minutes': 22
            },
            'performance': {
                'avg_latency_ms': 2.5,
                'p99_latency_ms': 10.2,
                'orders_per_second': 150,
                'data_processed_gb': 850
            },
            'errors': {
                'total': 42,
                'by_type': {
                    'connection': 15,
                    'timeout': 10,
                    'validation': 17
                }
            },
            'resources': {
                'cpu_usage_avg': 45.5,
                'memory_usage_avg': 62.3,
                'disk_usage': 75.8,
                'network_bandwidth_mbps': 250
            }
        }
    
    async def _collect_market_data(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Collecte les données de marché"""
        # Simuler pour la démo
        days = (end_date - start_date).days
        dates = pd.date_range(start=start_date, end=end_date, freq='D')
        
        # Prix simulés
        btc_prices = 50000 + np.cumsum(np.random.normal(0, 500, days))
        eth_prices = 3000 + np.cumsum(np.random.normal(0, 50, days))
        
        return {
            'prices': {
                'BTC-USDT': {
                    'dates': dates.tolist(),
                    'prices': btc_prices.tolist(),
                    'change_pct': float((btc_prices[-1] - btc_prices[0]) / btc_prices[0] * 100)
                },
                'ETH-USDT': {
                    'dates': dates.tolist(),
                    'prices': eth_prices.tolist(),
                    'change_pct': float((eth_prices[-1] - eth_prices[0]) / eth_prices[0] * 100)
                }
            },
            'volume': {
                'total_24h': 1500000000,
                'avg_daily': 1200000000
            },
            'volatility': {
                'BTC-USDT': 0.65,
                'ETH-USDT': 0.85
            }
        }
    
    async def _generate_pdf_report(self, config: ReportConfig, data: Dict[str, Any]) -> Path:
        """Génère un rapport PDF"""
        logger.info("Generating PDF report")
        
        # Nom du fichier
        filename = f"{config.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        report_path = self.reports_dir / filename
        
        # Créer le document
        doc = SimpleDocTemplate(
            str(report_path),
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Contenu du rapport
        story = []
        
        # Page de titre
        story.append(Paragraph(
            f"{config.name}",
            self.pdf_styles['CustomTitle']
        ))
        
        story.append(Spacer(1, 0.2 * inch))
        
        story.append(Paragraph(
            f"Period: {data['metadata']['period']['start'][:10]} to {data['metadata']['period']['end'][:10]}",
            self.pdf_styles['Normal']
        ))
        
        story.append(Spacer(1, 0.5 * inch))
        
        # Executive Summary
        if 'performance' in data:
            story.append(Paragraph("Executive Summary", self.pdf_styles['SectionTitle']))
            story.append(self._create_summary_table(data['performance']['summary']))
            story.append(PageBreak())
        
        # Performance Section
        if 'performance' in data:
            story.append(Paragraph("Performance Analysis", self.pdf_styles['SectionTitle']))
            
            # Graphique de P&L
            pnl_chart = self._create_pnl_chart(data['performance']['daily_pnl'])
            if pnl_chart:
                story.append(pnl_chart)
                story.append(Spacer(1, 0.3 * inch))
            
            # Métriques détaillées
            metrics_table = self._create_metrics_table(data['performance']['metrics'])
            story.append(metrics_table)
            story.append(PageBreak())
        
        # Trading Activity
        if 'trading' in data:
            story.append(Paragraph("Trading Activity", self.pdf_styles['SectionTitle']))
            
            # Résumé des trades
            trades_summary = self._create_trades_summary_table(data['trading'])
            story.append(trades_summary)
            story.append(Spacer(1, 0.3 * inch))
            
            # Performance par stratégie
            strategy_table = self._create_strategy_performance_table(data['trading']['strategy_performance'])
            story.append(strategy_table)
            story.append(PageBreak())
        
        # Risk Analysis
        if 'risk' in data:
            story.append(Paragraph("Risk Analysis", self.pdf_styles['SectionTitle']))
            
            risk_table = self._create_risk_table(data['risk'])
            story.append(risk_table)
            story.append(PageBreak())
        
        # System Performance
        if 'system' in data:
            story.append(Paragraph("System Performance", self.pdf_styles['SectionTitle']))
            
            system_table = self._create_system_table(data['system'])
            story.append(system_table)
        
        # Générer le PDF
        doc.build(story)
        
        logger.info(f"PDF report generated: {report_path}")
        return report_path
    
    def _create_summary_table(self, summary: Dict[str, Any]) -> Table:
        """Crée le tableau de résumé"""
        data = [
            ['Metric', 'Value'],
            ['Total Return', f"{summary['total_return']:.2%}"],
            ['Total P&L', f"${summary['total_pnl']:,.2f}"],
            ['Sharpe Ratio', f"{summary['sharpe_ratio']:.2f}"],
            ['Max Drawdown', f"{summary['max_drawdown']:.2%}"],
            ['Win Rate', f"{summary['win_rate']:.2%}"],
            ['Total Trades', f"{summary['total_trades']:,}"]
        ]
        
        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_pnl_chart(self, pnl_data: Dict[str, Any]) -> Optional[Image]:
        """Crée le graphique de P&L"""
        try:
            # Créer le graphique avec matplotlib
            fig, ax = plt.subplots(figsize=(8, 4))
            
            dates = pd.to_datetime(pnl_data['dates'])
            cumulative_pnl = pnl_data['cumulative_pnl']
            
            ax.plot(dates, cumulative_pnl, linewidth=2, color=self.colors['primary'])
            ax.fill_between(dates, cumulative_pnl, alpha=0.3, color=self.colors['primary'])
            
            ax.set_title('Cumulative P&L', fontsize=14, fontweight='bold')
            ax.set_xlabel('Date')
            ax.set_ylabel('P&L ($)')
            ax.grid(True, alpha=0.3)
            
            # Formatter les dates
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            ax.xaxis.set_major_locator(mdates.DayLocator(interval=7))
            plt.xticks(rotation=45)
            
            # Sauvegarder en mémoire
            img_buffer = io.BytesIO()
            plt.tight_layout()
            plt.savefig(img_buffer, format='png', dpi=150)
            plt.close()
            
            img_buffer.seek(0)
            
            # Créer l'image pour ReportLab
            return Image(img_buffer, width=6 * inch, height=3 * inch)
            
        except Exception as e:
            logger.error(f"Error creating P&L chart: {str(e)}")
            return None
    
    def _create_metrics_table(self, metrics: PerformanceMetrics) -> Table:
        """Crée le tableau des métriques détaillées"""
        data = [
            ['Risk Metrics', '', 'Trading Metrics', ''],
            ['Sharpe Ratio', f"{metrics.sharpe_ratio:.2f}", 'Win Rate', f"{metrics.win_rate:.2%}"],
            ['Sortino Ratio', f"{metrics.sortino_ratio:.2f}", 'Profit Factor', f"{metrics.profit_factor:.2f}"],
            ['Calmar Ratio', f"{metrics.calmar_ratio:.2f}", 'Expectancy', f"${metrics.expectancy:.2f}"],
            ['Max Drawdown', f"{metrics.max_drawdown:.2%}", 'Avg Win', f"${metrics.avg_win:.2f}"],
            ['Volatility', f"{metrics.volatility:.2%}", 'Avg Loss', f"${metrics.avg_loss:.2f}"],
            ['Beta', f"{metrics.beta:.2f}", 'Payoff Ratio', f"{metrics.payoff_ratio:.2f}"]
        ]
        
        table = Table(data, colWidths=[2 * inch, 1.5 * inch, 2 * inch, 1.5 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['dark'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        return table
    
    def _create_trades_summary_table(self, trading_data: Dict[str, Any]) -> Table:
        """Crée le tableau résumé des trades"""
        summary = trading_data['summary']
        
        data = [
            ['Trading Summary', 'Value'],
            ['Total Trades', f"{summary['total_trades']:,}"],
            ['Total Volume', f"${summary['total_volume']:,.2f}"],
            ['Avg Trade Size', f"${summary['avg_trade_size']:,.2f}"],
            ['Total Commission', f"${summary['total_commission']:,.2f}"]
        ]
        
        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['secondary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_strategy_performance_table(self, strategy_perf: Dict[str, Any]) -> Table:
        """Crée le tableau de performance par stratégie"""
        data = [['Strategy', 'Trades', 'Total P&L', 'Avg P&L', 'Win Rate']]
        
        for strategy, perf in strategy_perf.items():
            data.append([
                strategy,
                f"{perf['trades_count']:,}",
                f"${perf['total_pnl']:,.2f}",
                f"${perf['avg_pnl']:.2f}",
                f"{perf['win_rate']:.2%}"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['info'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        return table
    
    def _create_risk_table(self, risk_data: Dict[str, Any]) -> Table:
        """Crée le tableau d'analyse des risques"""
        data = [
            ['Risk Metric', 'Value'],
            ['VaR 95%', f"${risk_data['var']['95']:,.2f}"],
            ['VaR 99%', f"${risk_data['var']['99']:,.2f}"],
            ['Current Leverage', f"{risk_data['leverage']['current']:.1f}x"],
            ['Max Leverage', f"{risk_data['leverage']['maximum']:.1f}x"]
        ]
        
        # Ajouter les stress tests
        for test, value in risk_data['stress_tests'].items():
            data.append([test.replace('_', ' ').title(), f"${value:,.2f}"])
        
        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['danger'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_system_table(self, system_data: Dict[str, Any]) -> Table:
        """Crée le tableau de performance système"""
        data = [
            ['System Metric', 'Value'],
            ['Uptime', f"{system_data['uptime']['percentage']:.2f}%"],
            ['Avg Latency', f"{system_data['performance']['avg_latency_ms']:.1f} ms"],
            ['P99 Latency', f"{system_data['performance']['p99_latency_ms']:.1f} ms"],
            ['Orders/Second', f"{system_data['performance']['orders_per_second']:,}"],
            ['CPU Usage', f"{system_data['resources']['cpu_usage_avg']:.1f}%"],
            ['Memory Usage', f"{system_data['resources']['memory_usage_avg']:.1f}%"],
            ['Total Errors', f"{system_data['errors']['total']:,}"]
        ]
        
        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['success'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    async def _generate_html_report(self, config: ReportConfig, data: Dict[str, Any]) -> Path:
        """Génère un rapport HTML interactif"""
        logger.info("Generating HTML report")
        
        filename = f"{config.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        report_path = self.reports_dir / filename
        
        # Template HTML
        html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; }
        .metric-card { background: #f8f9fa; border-radius: 10px; padding: 20px; margin: 10px 0; }
        .metric-value { font-size: 2rem; font-weight: bold; color: #1f77b4; }
        .metric-label { color: #6c757d; font-size: 0.9rem; }
        .positive { color: #28a745; }
        .negative { color: #dc3545; }
        .chart-container { height: 400px; margin: 20px 0; }
        .table-container { margin: 20px 0; }
        h1, h2, h3 { color: #343a40; }
        .navbar { background-color: #1f77b4 !important; }
        .section { margin: 40px 0; }
    </style>
</head>
<body>
    <nav class="navbar navbar-dark">
        <div class="container-fluid">
            <span class="navbar-brand mb-0 h1">AI Trading Robot - {{ title }}</span>
            <span class="navbar-text">
                Generated: {{ generated_at }}
            </span>
        </div>
    </nav>
    
    <div class="container-fluid mt-4">
        <div class="row">
            <div class="col-md-12">
                <h1>{{ title }}</h1>
                <p class="lead">Period: {{ period_start }} to {{ period_end }}</p>
            </div>
        </div>
        
        <!-- Executive Summary -->
        <div class="section">
            <h2>Executive Summary</h2>
            <div class="row">
                {{ summary_cards }}
            </div>
        </div>
        
        <!-- Performance Charts -->
        <div class="section">
            <h2>Performance Analysis</h2>
            <div class="row">
                <div class="col-md-12">
                    <div id="pnl-chart" class="chart-container"></div>
                </div>
            </div>
            <div class="row">
                <div class="col-md-6">
                    <div id="returns-dist" class="chart-container"></div>
                </div>
                <div class="col-md-6">
                    <div id="drawdown-chart" class="chart-container"></div>
                </div>
            </div>
        </div>
        
        <!-- Trading Activity -->
        <div class="section">
            <h2>Trading Activity</h2>
            <div class="row">
                <div class="col-md-6">
                    <div id="strategy-performance" class="chart-container"></div>
                </div>
                <div class="col-md-6">
                    <div id="symbol-distribution" class="chart-container"></div>
                </div>
            </div>
            {{ trades_table }}
        </div>
        
        <!-- Risk Analysis -->
        <div class="section">
            <h2>Risk Analysis</h2>
            <div class="row">
                <div class="col-md-6">
                    <div id="exposure-chart" class="chart-container"></div>
                </div>
                <div class="col-md-6">
                    <div id="var-chart" class="chart-container"></div>
                </div>
            </div>
        </div>
        
        <!-- System Performance -->
        <div class="section">
            <h2>System Performance</h2>
            <div class="row">
                <div class="col-md-12">
                    {{ system_metrics }}
                </div>
            </div>
        </div>
    </div>
    
    <script>
        {{ chart_scripts }}
    </script>
</body>
</html>
"""
        
        # Préparer les données pour le template
        template_data = {
            'title': config.name,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'period_start': data['metadata']['period']['start'][:10],
            'period_end': data['metadata']['period']['end'][:10],
            'summary_cards': self._generate_summary_cards_html(data),
            'trades_table': self._generate_trades_table_html(data),
            'system_metrics': self._generate_system_metrics_html(data),
            'chart_scripts': self._generate_chart_scripts(data)
        }
        
        # Remplacer les variables dans le template
        html_content = html_template
        for key, value in template_data.items():
            html_content = html_content.replace(f'{{{{ {key} }}}}', str(value))
        
        # Sauvegarder le fichier
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"HTML report generated: {report_path}")
        return report_path
    
    def _generate_summary_cards_html(self, data: Dict[str, Any]) -> str:
        """Génère les cartes de résumé HTML"""
        if 'performance' not in data:
            return ""
        
        summary = data['performance']['summary']
        
        cards = []
        metrics = [
            ('Total Return', f"{summary['total_return']:.2%}", summary['total_return'] >= 0),
            ('Total P&L', f"${summary['total_pnl']:,.2f}", summary['total_pnl'] >= 0),
            ('Sharpe Ratio', f"{summary['sharpe_ratio']:.2f}", summary['sharpe_ratio'] >= 1),
            ('Max Drawdown', f"{summary['max_drawdown']:.2%}", True),
            ('Win Rate', f"{summary['win_rate']:.2%}", summary['win_rate'] >= 0.5),
            ('Total Trades', f"{summary['total_trades']:,}", True)
        ]
        
        for label, value, is_positive in metrics:
            color_class = 'positive' if is_positive else 'negative'
            cards.append(f"""
                <div class="col-md-2">
                    <div class="metric-card text-center">
                        <div class="metric-label">{label}</div>
                        <div class="metric-value {color_class}">{value}</div>
                    </div>
                </div>
            """)
        
        return ''.join(cards)
    
    def _generate_trades_table_html(self, data: Dict[str, Any]) -> str:
        """Génère le tableau des trades HTML"""
        if 'trading' not in data or 'trades' not in data['trading']:
            return ""
        
        trades = data['trading']['trades'][:10]  # Top 10 trades
        
        rows = []
        for trade in trades:
            pnl_class = 'positive' if trade['pnl'] >= 0 else 'negative'
            rows.append(f"""
                <tr>
                    <td>{trade['timestamp'][:16]}</td>
                    <td>{trade['symbol']}</td>
                    <td>{trade['strategy']}</td>
                    <td>{trade['side']}</td>
                    <td>{trade['quantity']:.4f}</td>
                    <td>${trade['price']:,.2f}</td>
                    <td class="{pnl_class}">${trade['pnl']:,.2f}</td>
                </tr>
            """)
        
        return f"""
        <div class="table-container">
            <h3>Recent Trades</h3>
            <table class="table table-striped">
                <thead>
                    <tr>
                        <th>Time</th>
                        <th>Symbol</th>
                        <th>Strategy</th>
                        <th>Side</th>
                        <th>Quantity</th>
                        <th>Price</th>
                        <th>P&L</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
        </div>
        """
    
    def _generate_system_metrics_html(self, data: Dict[str, Any]) -> str:
        """Génère les métriques système HTML"""
        if 'system' not in data:
            return ""
        
        system = data['system']
        
        return f"""
        <div class="row">
            <div class="col-md-3">
                <div class="metric-card">
                    <h4>Uptime</h4>
                    <div class="metric-value">{system['uptime']['percentage']:.2f}%</div>
                    <small class="text-muted">Total: {system['uptime']['total_hours']} hours</small>
                </div>
            </div>
            <div class="col-md-3">
                <div class="metric-card">
                    <h4>Performance</h4>
                    <div>Avg Latency: {system['performance']['avg_latency_ms']:.1f} ms</div>
                    <div>Orders/sec: {system['performance']['orders_per_second']:,}</div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="metric-card">
                    <h4>Resources</h4>
                    <div>CPU: {system['resources']['cpu_usage_avg']:.1f}%</div>
                    <div>Memory: {system['resources']['memory_usage_avg']:.1f}%</div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="metric-card">
                    <h4>Errors</h4>
                    <div class="metric-value negative">{system['errors']['total']}</div>
                    <small class="text-muted">Last 24 hours</small>
                </div>
            </div>
        </div>
        """
    
    def _generate_chart_scripts(self, data: Dict[str, Any]) -> str:
        """Génère les scripts Plotly pour les graphiques"""
        scripts = []
        
        # P&L Chart
        if 'performance' in data and 'daily_pnl' in data['performance']:
            pnl_data = data['performance']['daily_pnl']
            scripts.append(f"""
                // P&L Chart
                var pnlTrace = {{
                    x: {json.dumps([d[:10] for d in pnl_data['dates']])},
                    y: {json.dumps(pnl_data['cumulative_pnl'])},
                    type: 'scatter',
                    mode: 'lines',
                    fill: 'tozeroy',
                    name: 'Cumulative P&L'
                }};
                
                var pnlLayout = {{
                    title: 'Cumulative P&L Over Time',
                    xaxis: {{ title: 'Date' }},
                    yaxis: {{ title: 'P&L ($)' }},
                    showlegend: false
                }};
                
                Plotly.newPlot('pnl-chart', [pnlTrace], pnlLayout);
            """)
        
        # Strategy Performance
        if 'trading' in data and 'strategy_performance' in data['trading']:
            strategies = list(data['trading']['strategy_performance'].keys())
            pnls = [data['trading']['strategy_performance'][s]['total_pnl'] for s in strategies]
            
            scripts.append(f"""
                // Strategy Performance
                var strategyTrace = {{
                    x: {json.dumps(strategies)},
                    y: {json.dumps(pnls)},
                    type: 'bar',
                    marker: {{ color: ['#1f77b4', '#ff7f0e', '#2ca02c'] }}
                }};
                
                var strategyLayout = {{
                    title: 'Performance by Strategy',
                    xaxis: {{ title: 'Strategy' }},
                    yaxis: {{ title: 'Total P&L ($)' }}
                }};
                
                Plotly.newPlot('strategy-performance', [strategyTrace], strategyLayout);
            """)
        
        # Exposure Chart
        if 'risk' in data and 'exposure' in data['risk']:
            exposure = data['risk']['exposure']
            scripts.append(f"""
                // Exposure Chart
                var exposureTrace = {{
                    labels: {json.dumps(list(exposure.keys()))},
                    values: {json.dumps(list(exposure.values()))},
                    type: 'pie',
                    hole: 0.4
                }};
                
                var exposureLayout = {{
                    title: 'Portfolio Exposure'
                }};
                
                Plotly.newPlot('exposure-chart', [exposureTrace], exposureLayout);
            """)
        
        return '\n'.join(scripts)
    
    async def _generate_excel_report(self, config: ReportConfig, data: Dict[str, Any]) -> Path:
        """Génère un rapport Excel"""
        logger.info("Generating Excel report")
        
        filename = f"{config.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = self.reports_dir / filename
        
        # Créer le workbook
        wb = Workbook()
        
        # Feuille Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, data)
        
        # Feuille Performance
        if 'performance' in data:
            ws_perf = wb.create_sheet("Performance")
            self._create_performance_sheet(ws_perf, data['performance'])
        
        # Feuille Trading
        if 'trading' in data:
            ws_trading = wb.create_sheet("Trading")
            self._create_trading_sheet(ws_trading, data['trading'])
        
        # Feuille Risk
        if 'risk' in data:
            ws_risk = wb.create_sheet("Risk")
            self._create_risk_sheet(ws_risk, data['risk'])
        
        # Feuille System
        if 'system' in data:
            ws_system = wb.create_sheet("System")
            self._create_system_sheet(ws_system, data['system'])
        
        # Sauvegarder
        wb.save(report_path)
        
        logger.info(f"Excel report generated: {report_path}")
        return report_path
    
    def _create_summary_sheet(self, ws, data: Dict[str, Any]):
        """Crée la feuille de résumé Excel"""
        # Style du titre
        title_font = Font(size=16, bold=True, color="1f77b4")
        header_font = Font(size=12, bold=True)
        
        # Titre
        ws['A1'] = "AI Trading Robot - Performance Report"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        ws['A3'] = f"Period: {data['metadata']['period']['start'][:10]} to {data['metadata']['period']['end'][:10]}"
        
        # Métriques clés
        if 'performance' in data:
            summary = data['performance']['summary']
            
            row = 5
            ws[f'A{row}'] = "Key Metrics"
            ws[f'A{row}'].font = header_font
            ws.merge_cells(f'A{row}:B{row}')
            
            metrics = [
                ('Total Return', f"{summary['total_return']:.2%}"),
                ('Total P&L', f"${summary['total_pnl']:,.2f}"),
                ('Sharpe Ratio', f"{summary['sharpe_ratio']:.2f}"),
                ('Max Drawdown', f"{summary['max_drawdown']:.2%}"),
                ('Win Rate', f"{summary['win_rate']:.2%}"),
                ('Total Trades', f"{summary['total_trades']:,}")
            ]
            
            for i, (label, value) in enumerate(metrics):
                row = 7 + i
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                
                # Formater les cellules
                if '$' in value:
                    ws[f'B{row}'].alignment = Alignment(horizontal='right')
        
        # Ajuster les colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_performance_sheet(self, ws, performance_data: Dict[str, Any]):
        """Crée la feuille de performance Excel"""
        header_font = Font(bold=True)
        
        # Daily P&L data
        ws['A1'] = "Daily Performance"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Date', 'Daily P&L', 'Cumulative P&L']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        if 'daily_pnl' in performance_data:
            pnl_data = performance_data['daily_pnl']
            for i, date in enumerate(pnl_data['dates']):
                row = i + 4
                ws.cell(row=row, column=1, value=date[:10])
                ws.cell(row=row, column=2, value=pnl_data['pnl'][i])
                ws.cell(row=row, column=3, value=pnl_data['cumulative_pnl'][i])
        
        # Créer un graphique
        if len(pnl_data['dates']) > 0:
            chart = LineChart()
            chart.title = "Cumulative P&L"
            chart.y_axis.title = "P&L ($)"
            chart.x_axis.title = "Date"
            
            data_ref = Reference(ws, min_col=3, min_row=3, max_row=3+len(pnl_data['dates']))
            dates_ref = Reference(ws, min_col=1, min_row=4, max_row=3+len(pnl_data['dates']))
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(dates_ref)
            
            ws.add_chart(chart, "E3")
    
    def _create_trading_sheet(self, ws, trading_data: Dict[str, Any]):
        """Crée la feuille de trading Excel"""
        # À implémenter selon les besoins
        ws['A1'] = "Trading Activity"
        ws['A1'].font = Font(size=14, bold=True)
    
    def _create_risk_sheet(self, ws, risk_data: Dict[str, Any]):
        """Crée la feuille de risque Excel"""
        # À implémenter selon les besoins
        ws['A1'] = "Risk Analysis"
        ws['A1'].font = Font(size=14, bold=True)
    
    def _create_system_sheet(self, ws, system_data: Dict[str, Any]):
        """Crée la feuille système Excel"""
        # À implémenter selon les besoins
        ws['A1'] = "System Performance"
        ws['A1'].font = Font(size=14, bold=True)
    
    async def _generate_json_report(self, config: ReportConfig, data: Dict[str, Any]) -> Path:
        """Génère un rapport JSON"""
        logger.info("Generating JSON report")
        
        filename = f"{config.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        report_path = self.reports_dir / filename
        
        # Convertir les types non-sérialisables
        def serialize(obj):
            if isinstance(obj, (datetime, pd.Timestamp)):
                return obj.isoformat()
            elif isinstance(obj, Decimal):
                return float(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif hasattr(obj, '__dict__'):
                return obj.__dict__
            return str(obj)
        
        # Sauvegarder le JSON
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=serialize)
        
        logger.info(f"JSON report generated: {report_path}")
        return report_path
    
    async def _generate_dashboard(self, config: ReportConfig, data: Dict[str, Any]) -> Path:
        """Génère un dashboard interactif avec Dash"""
        logger.info("Generating interactive dashboard")
        
        # Pour la démo, sauvegarder les données et retourner un lien
        filename = f"dashboard_{config.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        dashboard_path = self.reports_dir / filename
        
        # Sauvegarder les données du dashboard
        with open(dashboard_path, 'w') as f:
            json.dump({
                'config': config.name,
                'data': data,
                'url': f"http://localhost:8050/dashboard/{filename}"
            }, f, indent=2)
        
        logger.info(f"Dashboard data saved: {dashboard_path}")
        return dashboard_path
    
    async def _deliver_report(self, config: ReportConfig, report_path: Path) -> None:
        """Livre le rapport selon les méthodes configurées"""
        logger.info(f"Delivering report via {[m.value for m in config.delivery_methods]}")
        
        for method in config.delivery_methods:
            try:
                if method == DeliveryMethod.EMAIL:
                    await self._send_email_report(config, report_path)
                elif method == DeliveryMethod.S3:
                    await self._upload_to_s3(config, report_path)
                elif method == DeliveryMethod.WEBHOOK:
                    await self._send_webhook(config, report_path)
                elif method == DeliveryMethod.LOCAL:
                    # Déjà sauvegardé localement
                    pass
                    
            except Exception as e:
                logger.error(f"Delivery failed for {method.value}: {str(e)}")
    
    async def _send_email_report(self, config: ReportConfig, report_path: Path) -> None:
        """Envoie le rapport par email"""
        # Configuration email
        smtp_config = self.config.get('smtp', {})
        
        if not smtp_config:
            logger.warning("SMTP configuration not found")
            return
        
        # Créer le message
        msg = MIMEMultipart()
        msg['From'] = smtp_config.get('from', 'trading-bot@example.com')
        msg['To'] = ', '.join(config.recipients)
        msg['Subject'] = f"Trading Report - {config.name} - {datetime.now().strftime('%Y-%m-%d')}"
        
        # Corps du message
        body = f"""
        Trading Report Generated
        
        Report: {config.name}
        Type: {config.type.value}
        Period: {datetime.now().strftime('%Y-%m-%d')}
        
        Please find the report attached.
        
        Best regards,
        AI Trading Robot
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attacher le fichier
        with open(report_path, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {report_path.name}'
            )
            msg.attach(part)
        
        # Envoyer l'email
        try:
            with smtplib.SMTP(smtp_config['host'], smtp_config['port']) as server:
                if smtp_config.get('tls', True):
                    server.starttls()
                if smtp_config.get('username'):
                    server.login(smtp_config['username'], smtp_config['password'])
                server.send_message(msg)
                
            logger.info(f"Email sent to {config.recipients}")
            
        except Exception as e:
            logger.error(f"Failed to send email: {str(e)}")
    
    async def _upload_to_s3(self, config: ReportConfig, report_path: Path) -> None:
        """Upload le rapport vers S3"""
        s3_config = self.config.get('s3', {})
        
        if not s3_config:
            logger.warning("S3 configuration not found")
            return
        
        try:
            s3_client = boto3.client(
                's3',
                aws_access_key_id=s3_config['access_key'],
                aws_secret_access_key=s3_config['secret_key']
            )
            
            bucket = s3_config['bucket']
            key = f"reports/{datetime.now().strftime('%Y/%m/%d')}/{report_path.name}"
            
            s3_client.upload_file(str(report_path), bucket, key)
            
            logger.info(f"Report uploaded to S3: s3://{bucket}/{key}")
            
        except Exception as e:
            logger.error(f"Failed to upload to S3: {str(e)}")
    
    async def _send_webhook(self, config: ReportConfig, report_path: Path) -> None:
        """Envoie une notification webhook"""
        webhook_url = config.custom_params.get('webhook_url')
        
        if not webhook_url:
            logger.warning("Webhook URL not configured")
            return
        
        try:
            # Préparer le payload
            payload = {
                'report_name': config.name,
                'report_type': config.type.value,
                'generated_at': datetime.now().isoformat(),
                'file_path': str(report_path),
                'file_size': report_path.stat().st_size
            }
            
            # Envoyer la requête
            response = requests.post(webhook_url, json=payload, timeout=30)
            response.raise_for_status()
            
            logger.info(f"Webhook sent to {webhook_url}")
            
        except Exception as e:
            logger.error(f"Failed to send webhook: {str(e)}")
    
    def _get_cache_key(self, config: ReportConfig) -> str:
        """Génère une clé de cache pour la configuration"""
        key_data = f"{config.name}_{config.type.value}_{json.dumps(config.filters, sort_keys=True)}"
        return hashlib.md5(key_data.encode()).hexdigest()


class ReportScheduler:
    """Planificateur de rapports automatiques"""
    
    def __init__(self, generator: ReportGenerator):
        self.generator = generator
        self.scheduled_reports: List[ReportConfig] = []
        self._running = False
        self._task = None
        
        logger.info("Report scheduler initialized")
    
    def add_scheduled_report(self, config: ReportConfig) -> None:
        """Ajoute un rapport planifié"""
        self.scheduled_reports.append(config)
        logger.info(f"Scheduled report added: {config.name} ({config.schedule})")
    
    async def start(self) -> None:
        """Démarre le planificateur"""
        self._running = True
        self._task = asyncio.create_task(self._run_scheduler())
        logger.info("Report scheduler started")
    
    async def stop(self) -> None:
        """Arrête le planificateur"""
        self._running = False
        if self._task:
            await self._task
        logger.info("Report scheduler stopped")
    
    async def _run_scheduler(self) -> None:
        """Boucle principale du planificateur"""
        while self._running:
            try:
                # Vérifier les rapports à générer
                now = datetime.now()
                
                for config in self.scheduled_reports:
                    if self._should_generate_report(config, now):
                        asyncio.create_task(self._generate_scheduled_report(config))
                
                # Attendre avant la prochaine vérification
                await asyncio.sleep(60)  # Vérifier chaque minute
                
            except Exception as e:
                logger.error(f"Scheduler error: {str(e)}", exc_info=True)
                await asyncio.sleep(60)
    
    def _should_generate_report(self, config: ReportConfig, now: datetime) -> bool:
        """Détermine si un rapport doit être généré"""
        if not config.schedule:
            return False
        
        # Logique simplifiée pour la démo
        # En production, utiliser croniter ou similar
        
        if config.type == ReportType.DAILY:
            # Générer à 6h du matin
            return now.hour == 6 and now.minute == 0
        elif config.type == ReportType.WEEKLY:
            # Générer le lundi à 6h
            return now.weekday() == 0 and now.hour == 6 and now.minute == 0
        elif config.type == ReportType.MONTHLY:
            # Générer le 1er du mois à 6h
            return now.day == 1 and now.hour == 6 and now.minute == 0
        
        return False
    
    async def _generate_scheduled_report(self, config: ReportConfig) -> None:
        """Génère un rapport planifié"""
        try:
            logger.info(f"Generating scheduled report: {config.name}")
            await self.generator.generate_report(config)
            
        except Exception as e:
            logger.error(f"Failed to generate scheduled report {config.name}: {str(e)}")


# Fonctions utilitaires pour une utilisation simplifiée

async def generate_daily_report() -> Path:
    """Génère le rapport quotidien standard"""
    generator = ReportGenerator()
    
    config = ReportConfig(
        name="Daily Trading Report",
        type=ReportType.DAILY,
        format=ReportFormat.PDF,
        delivery_methods=[DeliveryMethod.EMAIL, DeliveryMethod.S3],
        recipients=["trading-team@example.com"],
        include_sections=["performance", "trading", "risk", "system"]
    )
    
    return await generator.generate_report(config)


async def generate_performance_report(
    start_date: datetime,
    end_date: datetime,
    format: ReportFormat = ReportFormat.HTML
) -> Path:
    """Génère un rapport de performance personnalisé"""
    generator = ReportGenerator()
    
    config = ReportConfig(
        name="Performance Analysis",
        type=ReportType.ON_DEMAND,
        format=format,
        delivery_methods=[DeliveryMethod.LOCAL],
        filters={
            'start_date': start_date,
            'end_date': end_date
        },
        include_sections=["performance", "trading"]
    )
    
    return await generator.generate_report(config)


if __name__ == "__main__":
    # Test du module
    async def test_reporting():
        console.print("[bold]Testing Report Generation[/bold]")
        
        # Générer différents types de rapports
        generator = ReportGenerator()
        
        # Rapport PDF
        pdf_config = ReportConfig(
            name="Test Daily Report",
            type=ReportType.DAILY,
            format=ReportFormat.PDF,
            delivery_methods=[DeliveryMethod.LOCAL]
        )
        
        pdf_path = await generator.generate_report(pdf_config)
        console.print(f"✓ PDF Report: {pdf_path}")
        
        # Rapport HTML
        html_config = ReportConfig(
            name="Test Performance Report",
            type=ReportType.PERFORMANCE,
            format=ReportFormat.HTML,
            delivery_methods=[DeliveryMethod.LOCAL]
        )
        
        html_path = await generator.generate_report(html_config)
        console.print(f"✓ HTML Report: {html_path}")
        
        # Rapport Excel
        excel_config = ReportConfig(
            name="Test Trading Report",
            type=ReportType.WEEKLY,
            format=ReportFormat.EXCEL,
            delivery_methods=[DeliveryMethod.LOCAL]
        )
        
        excel_path = await generator.generate_report(excel_config)
        console.print(f"✓ Excel Report: {excel_path}")
        
        console.print("\n[green]All reports generated successfully![/green]")
    
    # Lancer le test
    asyncio.run(test_reporting())